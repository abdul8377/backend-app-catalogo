#!/usr/bin/env python3
"""Valida un paquete Excel v2 + ZIP de imágenes contra los datos maestros del backend."""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
import unicodedata
import uuid
import zipfile
from collections import defaultdict
from dataclasses import asdict, dataclass
from pathlib import Path, PurePosixPath
from typing import Any, Iterable

from openpyxl import load_workbook

REQUIRED_HEADERS = {
    "Fuentes": ["Lote", "ArchivoPDF", "Seccion", "PaginaDesde", "PaginaHasta", "Observacion"],
    "Productos": [
        "CodigoFamilia", "ProductoId", "Version", "Nombre", "Descripcion", "Empresa", "EmpresaId",
        "Marca", "MarcaId", "Categoria", "CategoriaId", "Subcategoria", "SubcategoriaId", "Tipo", "Estado",
    ],
    "Variantes": ["CodigoFamilia", "SKU", "CodigoProveedor", "NombreCorto", "Estado"],
    "Atributos": ["CodigoFamilia", "SKU", "Atributo", "Valor", "Unidad"],
    "Presentaciones": [
        "CodigoFamilia", "SKU", "Presentacion", "UnidadBase", "Equivalencia", "VentaMinima", "Incremento",
        "PermiteDecimales", "Estado",
    ],
    "Precios": [
        "CodigoFamilia", "SKU", "ListaPrecio", "Presentacion", "Moneda", "IGV", "Configuracion", "Precio", "Cotizar",
    ],
    "Imagenes": ["CodigoFamilia", "SKU", "Archivo", "Tipo", "Principal"],
}

REFERENCE_HEADERS = {
    "Ref_Empresas": ["EmpresaId", "Empresa"],
    "Ref_Marcas": ["MarcaId", "EmpresaId", "Marca"],
    "Ref_Categorias": ["CategoriaId", "CategoriaPadreId", "Categoria"],
    "Ref_Unidades": ["UnidadId", "Codigo", "Unidad", "Simbolo"],
    "Ref_Atributos": ["AtributoId", "CategoriaId", "Atributo", "Clave"],
}
OPTIONAL_REFERENCE_HEADERS = {
    "Ref_Marca_Categorias": ["RelacionId", "MarcaId", "CategoriaId"],
}

VALID_PRODUCT_TYPES = {"SINGLE", "LIST", "MATRIX"}
VALID_PRODUCT_STATUSES = {"DRAFT", "ACTIVE", "INACTIVE", "DELETED"}
VALID_ROW_STATUSES = {"ACTIVE", "INACTIVE"}
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}


@dataclass
class Message:
    severity: str
    sheet: str
    row: int
    column: str
    code: str
    message: str


class Validator:
    def __init__(
        self,
        workbook_path: Path,
        images_path: Path | None,
        references_path: Path | None,
        masters_path: Path | None = None,
    ):
        self.workbook_path = workbook_path
        self.images_path = images_path
        self.references_path = references_path
        self.masters_path = masters_path
        self.messages: list[Message] = []
        self.rows: dict[str, list[dict[str, Any]]] = {}
        self.zip_entries: set[str] = set()
        self.companies: dict[str, dict[str, Any]] = {}
        self.brands: dict[str, dict[str, Any]] = {}
        self.categories: dict[str, dict[str, Any]] = {}
        self.brand_categories: set[tuple[str, str]] | None = None
        self.units: set[str] = set()
        self.attributes_by_category: dict[str, set[str]] = defaultdict(set)
        self.reference_mode = references_path is not None

    def add(self, severity: str, sheet: str, row: int, column: str, code: str, message: str) -> None:
        self.messages.append(Message(severity, sheet, row, column, code, message))

    def run(self) -> dict[str, Any]:
        self._load_master_data()
        self._load_zip()
        self._load_workbook()
        self._validate_relations()
        self._validate_business_rules()
        counts = defaultdict(int)
        for msg in self.messages:
            counts[msg.severity] += 1
        return {
            "file": str(self.workbook_path),
            "images": str(self.images_path) if self.images_path else None,
            "references": str(self.references_path) if self.references_path else None,
            "masters": str(self.masters_path) if self.masters_path else None,
            "valid": counts["ERROR"] == 0,
            "counts": {
                "errors": counts["ERROR"],
                "warnings": counts["WARNING"],
                "info": counts["INFO"],
                "families": len(self.rows.get("Productos", [])),
                "variants": len(self.rows.get("Variantes", [])),
                "prices": len(self.rows.get("Precios", [])),
                "images": len(self.rows.get("Imagenes", [])),
            },
            "messages": [asdict(message) for message in self.messages],
        }

    def _load_master_data(self) -> None:
        if self.references_path:
            self._load_references_workbook()
            return
        if self.masters_path:
            self._load_legacy_masters()
            return
        self.add(
            "ERROR",
            "(referencias)",
            0,
            "",
            "REFERENCES_REQUIRED",
            "Debe proporcionar --references con referencias-datos-maestros.xlsx descargado del backend.",
        )

    def _load_references_workbook(self) -> None:
        path = self.references_path
        if not path or not path.exists():
            self.add("ERROR", "(referencias)", 0, "", "REFERENCES_NOT_FOUND", "El Excel de referencias no existe.")
            return
        if path.suffix.lower() != ".xlsx":
            self.add("ERROR", "(referencias)", 0, "", "INVALID_REFERENCES_EXTENSION", "Las referencias deben ser .xlsx.")
            return
        try:
            workbook = load_workbook(path, data_only=False, read_only=False)
        except Exception as exc:
            self.add("ERROR", "(referencias)", 0, "", "INVALID_REFERENCES", f"No se pudo abrir el Excel de referencias: {exc}")
            return

        reference_rows: dict[str, list[dict[str, Any]]] = {}
        for sheet_name, headers in REFERENCE_HEADERS.items():
            reference_rows[sheet_name] = self._read_reference_sheet(workbook, sheet_name, headers, required=True)
        for sheet_name, headers in OPTIONAL_REFERENCE_HEADERS.items():
            if sheet_name in workbook.sheetnames:
                reference_rows[sheet_name] = self._read_reference_sheet(workbook, sheet_name, headers, required=False)
            else:
                self.add(
                    "WARNING",
                    "(referencias)",
                    0,
                    sheet_name,
                    "BRAND_CATEGORY_REFERENCE_UNAVAILABLE",
                    "No existe Ref_Marca_Categorias; el backend deberá validar esta relación en la vista previa.",
                )
        workbook.close()

        for row in reference_rows.get("Ref_Empresas", []):
            self._register_reference(
                self.companies, row, "Ref_Empresas", "EmpresaId", "Empresa", "empresa",
                {"id": clean(row.get("EmpresaId")), "name": clean(row.get("Empresa"))},
            )

        for row in reference_rows.get("Ref_Marcas", []):
            self._register_reference(
                self.brands, row, "Ref_Marcas", "MarcaId", "Marca", "marca",
                {
                    "id": clean(row.get("MarcaId")),
                    "companyId": clean(row.get("EmpresaId")),
                    "name": clean(row.get("Marca")),
                },
            )

        for row in reference_rows.get("Ref_Categorias", []):
            self._register_reference(
                self.categories, row, "Ref_Categorias", "CategoriaId", "Categoria", "categoría",
                {
                    "id": clean(row.get("CategoriaId")),
                    "parentId": clean(row.get("CategoriaPadreId")),
                    "name": clean(row.get("Categoria")),
                    "level": upper(row.get("Nivel")),
                },
            )

        if "Ref_Marca_Categorias" in reference_rows:
            self.brand_categories = set()
            for row in reference_rows["Ref_Marca_Categorias"]:
                relation_id = clean(row.get("RelacionId"))
                brand_id = clean(row.get("MarcaId"))
                category_id = clean(row.get("CategoriaId"))
                self._require_reference_id(row, "Ref_Marca_Categorias", "RelacionId", relation_id)
                self._require_reference_id(row, "Ref_Marca_Categorias", "MarcaId", brand_id)
                self._require_reference_id(row, "Ref_Marca_Categorias", "CategoriaId", category_id)
                if brand_id and category_id:
                    self.brand_categories.add((brand_id, category_id))

        for row in reference_rows.get("Ref_Unidades", []):
            unit_id = clean(row.get("UnidadId"))
            self._require_reference_id(row, "Ref_Unidades", "UnidadId", unit_id)
            for column in ("Codigo", "Unidad", "Simbolo"):
                token = norm(row.get(column))
                if token:
                    self.units.add(token)

        for row in reference_rows.get("Ref_Atributos", []):
            attribute_id = clean(row.get("AtributoId"))
            category_id = clean(row.get("CategoriaId"))
            self._require_reference_id(row, "Ref_Atributos", "AtributoId", attribute_id)
            self._require_reference_id(row, "Ref_Atributos", "CategoriaId", category_id)
            if category_id and category_id not in self.categories:
                self.add(
                    "ERROR", "Ref_Atributos", row["_row"], "CategoriaId", "REFERENCE_ATTRIBUTE_CATEGORY_NOT_FOUND",
                    f"El atributo apunta a una categoría inexistente: {category_id}.",
                )
            for column in ("Atributo", "Clave"):
                token = norm(row.get(column))
                if category_id and token:
                    self.attributes_by_category[category_id].add(token)

        self._validate_reference_relations()

    def _read_reference_sheet(
        self,
        workbook: Any,
        sheet_name: str,
        required_headers: list[str],
        required: bool,
    ) -> list[dict[str, Any]]:
        if sheet_name not in workbook.sheetnames:
            if required:
                self.add("ERROR", sheet_name, 0, "", "MISSING_REFERENCE_SHEET", f"Falta la hoja {sheet_name}.")
            return []
        sheet = workbook[sheet_name]
        headers = [clean(cell.value) for cell in sheet[1]]
        missing = [header for header in required_headers if header not in headers]
        if missing:
            self.add(
                "ERROR", sheet_name, 1, "", "INVALID_REFERENCE_HEADERS",
                f"Faltan encabezados obligatorios: {missing}.",
            )
        parsed: list[dict[str, Any]] = []
        for row_index, cells in enumerate(sheet.iter_rows(min_row=2), start=2):
            values = [cell.value for cell in cells[: len(headers)]]
            if all(value is None or clean(value) == "" for value in values):
                continue
            for cell in cells:
                if cell.data_type == "f":
                    self.add("ERROR", sheet_name, row_index, cell.coordinate, "FORMULA_NOT_ALLOWED", "No se permiten fórmulas.")
            parsed.append({"_row": row_index, **dict(zip(headers, values))})
        return parsed

    def _register_reference(
        self,
        target: dict[str, dict[str, Any]],
        row: dict[str, Any],
        sheet: str,
        id_column: str,
        name_column: str,
        entity: str,
        value: dict[str, Any],
    ) -> None:
        item_id = clean(row.get(id_column))
        name = clean(row.get(name_column))
        self._require_reference_id(row, sheet, id_column, item_id)
        if not name:
            self.add("ERROR", sheet, row["_row"], name_column, "REFERENCE_NAME_REQUIRED", f"El nombre de {entity} es obligatorio.")
        if not item_id:
            return
        if item_id in target:
            self.add("ERROR", sheet, row["_row"], id_column, "DUPLICATE_REFERENCE_ID", f"ID de {entity} repetido: {item_id}.")
            return
        target[item_id] = value

    def _require_reference_id(self, row: dict[str, Any], sheet: str, column: str, item_id: str) -> None:
        if not item_id:
            self.add("ERROR", sheet, row["_row"], column, "REFERENCE_ID_REQUIRED", f"{column} es obligatorio.")
        elif not is_uuid(item_id):
            self.add("ERROR", sheet, row["_row"], column, "INVALID_REFERENCE_UUID", f"{column} no es un UUID válido: {item_id}.")

    def _validate_reference_relations(self) -> None:
        for brand in self.brands.values():
            if brand["companyId"] not in self.companies:
                self.add(
                    "ERROR", "Ref_Marcas", 0, "EmpresaId", "REFERENCE_BRAND_COMPANY_NOT_FOUND",
                    f"La marca {brand['id']} apunta a una empresa inexistente.",
                )
        for category in self.categories.values():
            parent_id = category["parentId"]
            if parent_id and parent_id not in self.categories:
                self.add(
                    "ERROR", "Ref_Categorias", 0, "CategoriaPadreId", "REFERENCE_PARENT_CATEGORY_NOT_FOUND",
                    f"La categoría {category['id']} apunta a una categoría padre inexistente.",
                )
        if self.brand_categories is not None:
            for brand_id, category_id in self.brand_categories:
                if brand_id not in self.brands:
                    self.add("ERROR", "Ref_Marca_Categorias", 0, "MarcaId", "REFERENCE_BRAND_NOT_FOUND", f"Marca inexistente: {brand_id}.")
                if category_id not in self.categories:
                    self.add("ERROR", "Ref_Marca_Categorias", 0, "CategoriaId", "REFERENCE_CATEGORY_NOT_FOUND", f"Categoría inexistente: {category_id}.")

    def _load_legacy_masters(self) -> None:
        path = self.masters_path
        try:
            data = json.loads(path.read_text(encoding="utf-8")) if path else {}
        except Exception as exc:
            self.add("ERROR", "(maestros)", 0, "", "INVALID_MASTERS", f"No se pudo leer datos maestros: {exc}")
            return
        self.add(
            "WARNING", "(maestros)", 0, "", "LEGACY_MASTERS_MODE",
            "--masters se conserva solo para pruebas; las importaciones reales deben usar --references.",
        )
        self.companies = {clean(item.get("id")): item for item in data.get("companies", []) if clean(item.get("id"))}
        self.brands = {clean(item.get("id")): item for item in data.get("brands", []) if clean(item.get("id"))}
        self.categories = {clean(item.get("id")): item for item in data.get("categories", []) if clean(item.get("id"))}
        if "brandCategoryRelations" in data:
            self.brand_categories = {
                (clean(item.get("brandId")), clean(item.get("categoryId")))
                for item in data.get("brandCategoryRelations", [])
                if clean(item.get("brandId")) and clean(item.get("categoryId"))
            }
        for item in data.get("categoryAttributes", []):
            category_id = clean(item.get("categoryId"))
            name = norm(item.get("name"))
            if category_id and name:
                self.attributes_by_category[category_id].add(name)

    def _load_zip(self) -> None:
        if not self.images_path:
            return
        if not self.images_path.exists():
            self.add("ERROR", "Imagenes", 0, "Archivo", "ZIP_NOT_FOUND", "El ZIP de imágenes no existe.")
            return
        try:
            with zipfile.ZipFile(self.images_path) as archive:
                for name in archive.namelist():
                    normalized = normalize_zip_path(name)
                    if normalized:
                        self.zip_entries.add(normalized)
        except zipfile.BadZipFile:
            self.add("ERROR", "Imagenes", 0, "Archivo", "INVALID_ZIP", "El archivo de imágenes no es un ZIP válido.")

    def _load_workbook(self) -> None:
        if not self.workbook_path.exists():
            self.add("ERROR", "(archivo)", 0, "", "XLSX_NOT_FOUND", "El archivo XLSX no existe.")
            return
        if self.workbook_path.suffix.lower() != ".xlsx":
            self.add("ERROR", "(archivo)", 0, "", "INVALID_EXTENSION", "Solo se admite .xlsx.")
            return
        try:
            workbook = load_workbook(self.workbook_path, data_only=False, read_only=False)
        except Exception as exc:
            self.add("ERROR", "(archivo)", 0, "", "INVALID_XLSX", f"No se pudo abrir el XLSX: {exc}")
            return

        for required_sheet, headers in REQUIRED_HEADERS.items():
            if required_sheet not in workbook.sheetnames:
                self.add("ERROR", required_sheet, 0, "", "MISSING_SHEET", f"Falta la hoja {required_sheet}.")
                self.rows[required_sheet] = []
                continue
            sheet = workbook[required_sheet]
            actual_headers = [clean(cell.value) for cell in sheet[1]][: len(headers)]
            if actual_headers != headers:
                self.add(
                    "ERROR", required_sheet, 1, "", "INVALID_HEADERS",
                    f"Encabezados esperados: {headers}. Encontrados: {actual_headers}.",
                )
            parsed: list[dict[str, Any]] = []
            for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                values = [cell.value for cell in row[: len(headers)]]
                if all(value is None or clean(value) == "" for value in values):
                    continue
                for cell in row:
                    if cell.data_type == "f":
                        self.add("ERROR", required_sheet, row_index, cell.coordinate, "FORMULA_NOT_ALLOWED", "No se permiten fórmulas.")
                parsed.append({"_row": row_index, **dict(zip(headers, values))})
            self.rows[required_sheet] = parsed
        workbook.close()

    def _validate_relations(self) -> None:
        products = self.rows.get("Productos", [])
        variants = self.rows.get("Variantes", [])
        family_rows: dict[str, dict[str, Any]] = {}
        for row in products:
            family = upper(row.get("CodigoFamilia"))
            if not family:
                self.add("ERROR", "Productos", row["_row"], "CodigoFamilia", "FAMILY_REQUIRED", "CodigoFamilia es obligatorio.")
                continue
            if family in family_rows:
                self.add("ERROR", "Productos", row["_row"], "CodigoFamilia", "DUPLICATE_FAMILY", f"Familia repetida: {family}.")
            family_rows[family] = row

        variants_by_family: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
        global_skus: dict[str, tuple[str, int]] = {}
        for row in variants:
            family = upper(row.get("CodigoFamilia"))
            sku = upper(row.get("SKU"))
            if family not in family_rows:
                self.add("ERROR", "Variantes", row["_row"], "CodigoFamilia", "UNKNOWN_FAMILY", f"La familia {family} no existe en Productos.")
            if not sku:
                self.add("ERROR", "Variantes", row["_row"], "SKU", "SKU_REQUIRED", "SKU es obligatorio.")
                continue
            if sku in variants_by_family[family]:
                self.add("ERROR", "Variantes", row["_row"], "SKU", "DUPLICATE_SKU_IN_FAMILY", f"SKU repetido en la familia: {sku}.")
            if sku in global_skus:
                other_family, other_row = global_skus[sku]
                self.add("ERROR", "Variantes", row["_row"], "SKU", "DUPLICATE_SKU_GLOBAL", f"SKU {sku} ya aparece en {other_family}, fila {other_row}.")
            variants_by_family[family][sku] = row
            global_skus[sku] = (family, row["_row"])

        for sheet_name in ("Atributos", "Presentaciones", "Precios", "Imagenes"):
            for row in self.rows.get(sheet_name, []):
                family = upper(row.get("CodigoFamilia"))
                sku = upper(row.get("SKU"))
                if family not in family_rows:
                    self.add("ERROR", sheet_name, row["_row"], "CodigoFamilia", "UNKNOWN_FAMILY", f"La familia {family} no existe.")
                if sku and sku not in variants_by_family.get(family, {}):
                    self.add("ERROR", sheet_name, row["_row"], "SKU", "UNKNOWN_SKU", f"El SKU {sku} no existe en la familia {family}.")

    def _validate_business_rules(self) -> None:
        products = self.rows.get("Productos", [])
        variants = self.rows.get("Variantes", [])
        attributes = self.rows.get("Atributos", [])
        presentations = self.rows.get("Presentaciones", [])
        prices = self.rows.get("Precios", [])
        images = self.rows.get("Imagenes", [])

        variants_by_family = group_by_family(variants)
        presentations_by_family = group_by_family(presentations)
        prices_by_family = group_by_family(prices)
        images_by_family = group_by_family(images)
        products_by_family = {upper(row.get("CodigoFamilia")): row for row in products}

        for row in products:
            family = upper(row.get("CodigoFamilia"))
            required_text(row, "Productos", self, ["Nombre", "Empresa", "Marca", "Categoria", "Tipo", "Estado"])
            product_type = upper(row.get("Tipo")) or "SINGLE"
            status = upper(row.get("Estado")) or "DRAFT"
            if product_type not in VALID_PRODUCT_TYPES:
                self.add("ERROR", "Productos", row["_row"], "Tipo", "INVALID_PRODUCT_TYPE", f"Tipo inválido: {product_type}.")
            if status not in VALID_PRODUCT_STATUSES:
                self.add("ERROR", "Productos", row["_row"], "Estado", "INVALID_PRODUCT_STATUS", f"Estado inválido: {status}.")
            family_variants = variants_by_family.get(family, [])
            if not family_variants:
                self.add("ERROR", "Productos", row["_row"], "CodigoFamilia", "NO_VARIANTS", "La familia no tiene variantes.")
            if product_type == "SINGLE" and len(family_variants) != 1:
                self.add("ERROR", "Productos", row["_row"], "Tipo", "SINGLE_VARIANT_COUNT", "SINGLE requiere exactamente una variante.")
            if not presentations_by_family.get(family):
                self.add("ERROR", "Productos", row["_row"], "CodigoFamilia", "NO_PRESENTATIONS", "La familia no tiene presentaciones.")
            if not images_by_family.get(family):
                self.add("WARNING", "Productos", row["_row"], "CodigoFamilia", "NO_IMAGES", "La familia no tiene imágenes.")
            if status == "ACTIVE":
                if not images_by_family.get(family):
                    self.add("ERROR", "Productos", row["_row"], "Estado", "ACTIVE_WITHOUT_IMAGE", "Un producto ACTIVE debe tener imagen.")
                if not prices_by_family.get(family):
                    self.add("ERROR", "Productos", row["_row"], "Estado", "ACTIVE_WITHOUT_PRICES", "Un producto ACTIVE debe tener precios o por cotizar.")
            self._validate_master_row(row)

        for row in variants:
            required_text(row, "Variantes", self, ["CodigoFamilia", "SKU", "NombreCorto", "Estado"])
            status = upper(row.get("Estado")) or "ACTIVE"
            if status not in VALID_ROW_STATUSES:
                self.add("ERROR", "Variantes", row["_row"], "Estado", "INVALID_VARIANT_STATUS", f"Estado inválido: {status}.")

        for row in presentations:
            required_text(row, "Presentaciones", self, ["CodigoFamilia", "Presentacion", "UnidadBase", "Estado"])
            for column in ("Equivalencia", "VentaMinima", "Incremento"):
                value = number(row.get(column))
                if value is None or value <= 0:
                    self.add("ERROR", "Presentaciones", row["_row"], column, "INVALID_POSITIVE_NUMBER", f"{column} debe ser mayor que cero.")
            unit = norm(row.get("UnidadBase"))
            if self.reference_mode and unit and unit not in self.units:
                self.add("ERROR", "Presentaciones", row["_row"], "UnidadBase", "UNKNOWN_MEASUREMENT_UNIT", "UnidadBase no existe en Ref_Unidades.")

        presentation_keys = {
            (upper(row.get("CodigoFamilia")), upper(row.get("SKU")), norm(row.get("Presentacion")))
            for row in presentations
        }
        for row in prices:
            family = upper(row.get("CodigoFamilia"))
            sku = upper(row.get("SKU"))
            presentation = norm(row.get("Presentacion"))
            required_text(row, "Precios", self, ["CodigoFamilia", "SKU", "ListaPrecio", "Presentacion", "Moneda", "IGV"])
            currency = upper(row.get("Moneda"))
            if currency != "PEN":
                self.add("ERROR", "Precios", row["_row"], "Moneda", "CURRENCY_MUST_BE_PEN", "La moneda debe ser PEN.")
            igv = number(row.get("IGV"))
            if igv is None:
                self.add("ERROR", "Precios", row["_row"], "IGV", "IGV_REQUIRED", "IGV es obligatorio.")
            elif not math.isclose(igv, 18.0, abs_tol=0.001):
                self.add("ERROR", "Precios", row["_row"], "IGV", "IGV_MUST_BE_18", f"El IGV debe ser 18; se recibió {igv}.")
            config = upper(row.get("Configuracion")) or "PRECIO_FIJO"
            quote = yes(row.get("Cotizar")) or config in {"POR_COTIZAR", "QUOTE"}
            price = number(row.get("Precio"))
            if quote:
                if price not in (None, 0):
                    self.add("WARNING", "Precios", row["_row"], "Precio", "QUOTE_WITH_PRICE", "Por cotizar contiene un precio; se recomienda dejarlo vacío.")
            elif price is None:
                self.add("ERROR", "Precios", row["_row"], "Precio", "FIXED_PRICE_REQUIRED", "Precio fijo requiere Precio.")
            elif price == 0:
                self.add("ERROR", "Precios", row["_row"], "Precio", "ZERO_PRICE_NOT_ALLOWED", "Precio 0 requiere confirmación explícita de gratuidad.")
            elif price < 0:
                self.add("ERROR", "Precios", row["_row"], "Precio", "NEGATIVE_PRICE", "El precio no puede ser negativo.")
            if (family, sku, presentation) not in presentation_keys and (family, "", presentation) not in presentation_keys:
                self.add("ERROR", "Precios", row["_row"], "Presentacion", "PRICE_PRESENTATION_NOT_FOUND", "La presentación no corresponde al SKU ni a la familia.")

        self._validate_attribute_references(attributes, products_by_family)
        self._validate_images(images)

    def _validate_master_row(self, row: dict[str, Any]) -> None:
        company_id = self._required_product_master_id(row, "EmpresaId")
        brand_id = self._required_product_master_id(row, "MarcaId")
        category_id = self._required_product_master_id(row, "CategoriaId")
        subcategory_id = clean(row.get("SubcategoriaId"))
        subcategory_name = clean(row.get("Subcategoria"))

        company = self.companies.get(company_id) if company_id else None
        brand = self.brands.get(brand_id) if brand_id else None
        category = self.categories.get(category_id) if category_id else None

        if company_id and not company:
            self.add("ERROR", "Productos", row["_row"], "EmpresaId", "UNKNOWN_COMPANY_ID", "EmpresaId no existe en las referencias.")
        elif company and not same_name(company.get("name"), row.get("Empresa")):
            self.add("ERROR", "Productos", row["_row"], "Empresa", "COMPANY_NAME_MISMATCH", "Empresa no coincide con el nombre asociado a EmpresaId.")

        if brand_id and not brand:
            self.add("ERROR", "Productos", row["_row"], "MarcaId", "UNKNOWN_BRAND_ID", "MarcaId no existe en las referencias.")
        elif brand:
            if not same_name(brand.get("name"), row.get("Marca")):
                self.add("ERROR", "Productos", row["_row"], "Marca", "BRAND_NAME_MISMATCH", "Marca no coincide con el nombre asociado a MarcaId.")
            if company_id and clean(brand.get("companyId")) != company_id:
                self.add("ERROR", "Productos", row["_row"], "MarcaId", "BRAND_COMPANY_MISMATCH", "La marca no pertenece a la empresa seleccionada.")

        if category_id and not category:
            self.add("ERROR", "Productos", row["_row"], "CategoriaId", "UNKNOWN_CATEGORY_ID", "CategoriaId no existe en las referencias.")
        elif category:
            if not same_name(category.get("name"), row.get("Categoria")):
                self.add("ERROR", "Productos", row["_row"], "Categoria", "CATEGORY_NAME_MISMATCH", "Categoria no coincide con el nombre asociado a CategoriaId.")
            if clean(category.get("parentId")):
                self.add("ERROR", "Productos", row["_row"], "CategoriaId", "CATEGORY_MUST_BE_ROOT", "CategoriaId apunta a una subcategoría; use CategoriaId para el padre y SubcategoriaId para el hijo.")

        if subcategory_name and not subcategory_id:
            self.add("ERROR", "Productos", row["_row"], "SubcategoriaId", "SUBCATEGORY_ID_REQUIRED", "SubcategoriaId es obligatorio cuando se informa Subcategoria.")
        if subcategory_id and not subcategory_name:
            self.add("ERROR", "Productos", row["_row"], "Subcategoria", "SUBCATEGORY_NAME_REQUIRED", "Subcategoria es obligatoria cuando se informa SubcategoriaId.")
        if subcategory_id:
            if self.reference_mode and not is_uuid(subcategory_id):
                self.add("ERROR", "Productos", row["_row"], "SubcategoriaId", "INVALID_MASTER_UUID", "SubcategoriaId debe ser un UUID del backend.")
            subcategory = self.categories.get(subcategory_id)
            if not subcategory:
                self.add("ERROR", "Productos", row["_row"], "SubcategoriaId", "UNKNOWN_SUBCATEGORY_ID", "SubcategoriaId no existe en Ref_Categorias.")
            else:
                if not same_name(subcategory.get("name"), subcategory_name):
                    self.add("ERROR", "Productos", row["_row"], "Subcategoria", "SUBCATEGORY_NAME_MISMATCH", "Subcategoria no coincide con el nombre asociado a SubcategoriaId.")
                if category_id and clean(subcategory.get("parentId")) != category_id:
                    self.add("ERROR", "Productos", row["_row"], "SubcategoriaId", "SUBCATEGORY_PARENT_MISMATCH", "La subcategoría no pertenece a la categoría seleccionada mediante categoria_padre_id.")

        if self.brand_categories is not None and brand_id and category_id:
            if (brand_id, category_id) not in self.brand_categories:
                self.add("ERROR", "Productos", row["_row"], "CategoriaId", "BRAND_CATEGORY_MISMATCH", "La marca no está relacionada con la categoría seleccionada.")

    def _required_product_master_id(self, row: dict[str, Any], column: str) -> str:
        item_id = clean(row.get(column))
        if not item_id:
            self.add("ERROR", "Productos", row["_row"], column, "MASTER_ID_REQUIRED", f"{column} es obligatorio y debe provenir del Excel de referencias.")
        elif self.reference_mode and not is_uuid(item_id):
            self.add("ERROR", "Productos", row["_row"], column, "INVALID_MASTER_UUID", f"{column} debe ser un UUID del backend.")
        return item_id

    def _validate_attribute_references(
        self,
        rows: list[dict[str, Any]],
        products_by_family: dict[str, dict[str, Any]],
    ) -> None:
        warned_attributes: set[tuple[str, str]] = set()
        warned_units: set[str] = set()
        for row in rows:
            family = upper(row.get("CodigoFamilia"))
            attribute = norm(row.get("Atributo"))
            required_text(row, "Atributos", self, ["CodigoFamilia", "Atributo", "Valor"])
            product = products_by_family.get(family)
            if not product or not attribute or not self.attributes_by_category:
                continue
            classification_id = clean(product.get("SubcategoriaId")) or clean(product.get("CategoriaId"))
            configured = self.attributes_by_category.get(classification_id, set())
            key = (classification_id, attribute)
            if attribute not in configured and key not in warned_attributes:
                warned_attributes.add(key)
                self.add(
                    "WARNING", "Atributos", row["_row"], "Atributo", "UNCONFIGURED_ATTRIBUTE",
                    "El atributo no está configurado para la categoría o subcategoría; inclúyalo como decisión pendiente.",
                )
            unit = norm(row.get("Unidad"))
            if self.reference_mode and unit and unit not in self.units and unit not in warned_units:
                warned_units.add(unit)
                self.add("WARNING", "Atributos", row["_row"], "Unidad", "UNKNOWN_ATTRIBUTE_UNIT", "La unidad del atributo no aparece en Ref_Unidades.")

    def _validate_images(self, images: list[dict[str, Any]]) -> None:
        principals: dict[tuple[str, str], int] = defaultdict(int)
        for row in images:
            family = upper(row.get("CodigoFamilia"))
            sku = upper(row.get("SKU"))
            file_name = clean(row.get("Archivo"))
            if not file_name:
                self.add("ERROR", "Imagenes", row["_row"], "Archivo", "IMAGE_FILE_REQUIRED", "Archivo es obligatorio.")
                continue
            normalized = normalize_zip_path(file_name)
            if not normalized or ".." in PurePosixPath(normalized).parts or PurePosixPath(normalized).is_absolute():
                self.add("ERROR", "Imagenes", row["_row"], "Archivo", "UNSAFE_IMAGE_PATH", "Ruta de imagen insegura.")
                continue
            if Path(normalized).suffix.lower() not in IMAGE_EXTENSIONS:
                self.add("ERROR", "Imagenes", row["_row"], "Archivo", "INVALID_IMAGE_EXTENSION", "Formato de imagen no permitido.")
            if self.images_path and normalized not in self.zip_entries:
                self.add("ERROR", "Imagenes", row["_row"], "Archivo", "IMAGE_NOT_IN_ZIP", f"{normalized} no existe dentro del ZIP.")
            if not self.images_path:
                self.add("ERROR", "Imagenes", row["_row"], "Archivo", "ZIP_REQUIRED", "Hay imágenes declaradas, pero no se proporcionó ZIP.")
            if yes(row.get("Principal")):
                principals[(family, sku)] += 1
        for (family, sku), count in principals.items():
            if count > 1:
                scope = f"SKU {sku}" if sku else "familia"
                self.add("ERROR", "Imagenes", 0, "Principal", "MULTIPLE_PRINCIPAL_IMAGES", f"La {scope} de {family} tiene {count} imágenes principales.")


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def upper(value: Any) -> str:
    return clean(value).upper()


def norm(value: Any) -> str:
    text = clean(value).casefold()
    text = "".join(char for char in unicodedata.normalize("NFKD", text) if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text).strip()


def same_name(left: Any, right: Any) -> bool:
    normalize = lambda value: re.sub(r"\s+", " ", clean(value)).strip().casefold()
    return normalize(left) == normalize(right)


def number(value: Any) -> float | None:
    text = clean(value).replace("S/", "").replace("$", "").replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def yes(value: Any) -> bool:
    return upper(value) in {"SI", "SÍ", "YES", "TRUE", "1", "X"}


def is_uuid(value: Any) -> bool:
    try:
        uuid.UUID(clean(value))
        return True
    except (ValueError, AttributeError):
        return False


def normalize_zip_path(value: str) -> str:
    raw = clean(value).replace("\\", "/")
    while raw.startswith("./"):
        raw = raw[2:]
    normalized = str(PurePosixPath(raw)) if raw else ""
    return "" if normalized == "." else normalized


def required_text(row: dict[str, Any], sheet: str, validator: Validator, columns: Iterable[str]) -> None:
    for column in columns:
        if not clean(row.get(column)):
            validator.add("ERROR", sheet, row["_row"], column, "REQUIRED_FIELD", f"{column} es obligatorio.")


def group_by_family(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[upper(row.get("CodigoFamilia"))].append(row)
    return grouped


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", required=True, type=Path)
    parser.add_argument("--images", type=Path)
    masters_group = parser.add_mutually_exclusive_group()
    masters_group.add_argument("--references", type=Path, help="Excel referencias-datos-maestros.xlsx descargado del backend.")
    masters_group.add_argument("--masters", type=Path, help=argparse.SUPPRESS)
    parser.add_argument("--report", type=Path, default=Path("reporte_validacion.json"))
    args = parser.parse_args()

    report = Validator(args.xlsx, args.images, args.references, args.masters).run()
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report["counts"], ensure_ascii=False, indent=2))
    print(f"Reporte: {args.report}")
    return 0 if report["valid"] else 1


if __name__ == "__main__":
    sys.exit(main())
